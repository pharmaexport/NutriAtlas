#!/usr/bin/env python3
"""
Import the official CIQUAL XLSX workbook into PostgreSQL.

Design goals:
- preserve every raw CIQUAL row as JSONB;
- normalize foods and nutrients for search/calculation;
- keep source version, checksum and worksheet metadata;
- never require committing the XLSX file to Git.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from psycopg.types.json import Jsonb

COMPOSITION_SHEET = "composition nutritionnelle"
INFOODS_SHEET = "codes INFOODS"
SOURCE_URL = "https://ciqual.anses.fr/"

FOOD_COLUMNS = {
    "alim_grp_code",
    "alim_ssgrp_code",
    "alim_ssssgrp_code",
    "alim_grp_nom_fr",
    "alim_ssgrp_nom_fr",
    "alim_ssssgrp_nom_fr",
    "alim_code",
    "alim_nom_fr",
    "alim_nom_sci",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def normalize_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    text = str(value).strip().replace(",", ".")
    if text in {"-", "traces", "<", ""}:
        return None
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def extract_unit(column_name: str) -> str | None:
    match = re.search(r"\(([^()]*)\)\s*$", column_name)
    if not match:
        return None
    return match.group(1).replace("100 g", "/100 g")


def load_infoods(workbook: openpyxl.Workbook) -> dict[str, dict[str, str | None]]:
    if INFOODS_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[INFOODS_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    mapping: dict[str, dict[str, str | None]] = {}
    for row in rows[1:]:
        if len(row) < 3:
            continue
        infoods_tag = clean(row[0]) or None
        origcpcd = clean(row[1]) or None
        const_name = clean(row[2])
        if const_name:
            mapping[const_name] = {
                "infoods_tag": infoods_tag,
                "origcpcd": origcpcd,
            }
    return mapping


def ensure_dataset_version(conn: psycopg.Connection, *, version: str, checksum: str, raw_file: Path) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO dataset_versions (
                dataset, publisher, version, source_url, source_worksheet,
                retrieved_at, checksum_sha256, raw_file_path
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (dataset, version, checksum_sha256)
            DO UPDATE SET raw_file_path = EXCLUDED.raw_file_path
            RETURNING id
            """,
            (
                "CIQUAL",
                "ANSES",
                version,
                SOURCE_URL,
                COMPOSITION_SHEET,
                datetime.now(timezone.utc),
                checksum,
                str(raw_file),
            ),
        )
        return int(cur.fetchone()[0])


def get_or_create_nutrient(
    conn: psycopg.Connection,
    *,
    source_column_name: str,
    infoods_tag: str | None,
    origcpcd: str | None,
    unit: str | None,
) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO nutrients (
                source_nutrient_code, infoods_tag, origcpcd, name, unit, source_column_name
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (source_column_name)
            DO UPDATE SET
                infoods_tag = EXCLUDED.infoods_tag,
                origcpcd = EXCLUDED.origcpcd,
                unit = EXCLUDED.unit
            RETURNING id
            """,
            (infoods_tag, infoods_tag, origcpcd, source_column_name, unit, source_column_name),
        )
        return int(cur.fetchone()[0])


def import_workbook(xlsx_path: Path, *, version: str, database_url: str) -> None:
    checksum = sha256_file(xlsx_path)
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if COMPOSITION_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {COMPOSITION_SHEET}")

    infoods = load_infoods(workbook)
    sheet = workbook[COMPOSITION_SHEET]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]

    with psycopg.connect(database_url) as conn:
        dataset_version_id = ensure_dataset_version(
            conn,
            version=version,
            checksum=checksum,
            raw_file=xlsx_path,
        )

        nutrient_ids: dict[str, int] = {}
        for header in headers:
            if header and header not in FOOD_COLUMNS:
                meta = infoods.get(header, {})
                nutrient_ids[header] = get_or_create_nutrient(
                    conn,
                    source_column_name=header,
                    infoods_tag=meta.get("infoods_tag"),
                    origcpcd=meta.get("origcpcd"),
                    unit=extract_unit(header),
                )

        imported = 0
        with conn.cursor() as cur:
            for source_row_number, row in enumerate(rows, start=2):
                raw = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
                source_food_code = clean(raw.get("alim_code"))
                if not source_food_code:
                    continue

                cur.execute(
                    """
                    INSERT INTO raw_food_rows (
                        dataset_version_id, source_row_number, source_food_code, raw_row_json
                    )
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (dataset_version_id, source_food_code)
                    DO UPDATE SET raw_row_json = EXCLUDED.raw_row_json
                    RETURNING id
                    """,
                    (dataset_version_id, source_row_number, source_food_code, Jsonb(raw)),
                )
                raw_food_row_id = int(cur.fetchone()[0])

                cur.execute(
                    """
                    INSERT INTO foods (
                        dataset_version_id, raw_food_row_id, source_food_code, name,
                        scientific_name, food_group_code, food_subgroup_code,
                        food_subsubgroup_code, food_group_name_fr, food_subgroup_name_fr,
                        food_subsubgroup_name_fr
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (dataset_version_id, source_food_code)
                    DO UPDATE SET
                        raw_food_row_id = EXCLUDED.raw_food_row_id,
                        name = EXCLUDED.name,
                        scientific_name = EXCLUDED.scientific_name
                    RETURNING id
                    """,
                    (
                        dataset_version_id,
                        raw_food_row_id,
                        source_food_code,
                        clean(raw.get("alim_nom_fr")),
                        clean(raw.get("alim_nom_sci")) or None,
                        clean(raw.get("alim_grp_code")) or None,
                        clean(raw.get("alim_ssgrp_code")) or None,
                        clean(raw.get("alim_ssssgrp_code")) or None,
                        clean(raw.get("alim_grp_nom_fr")) or None,
                        clean(raw.get("alim_ssgrp_nom_fr")) or None,
                        clean(raw.get("alim_ssssgrp_nom_fr")) or None,
                    ),
                )
                food_id = int(cur.fetchone()[0])

                for header, nutrient_id in nutrient_ids.items():
                    original_value = raw.get(header)
                    cur.execute(
                        """
                        INSERT INTO food_nutrients (
                            food_id, nutrient_id, value, original_value, source,
                            source_url, source_column_name, dataset_version_id
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (food_id, nutrient_id, dataset_version_id)
                        DO UPDATE SET
                            value = EXCLUDED.value,
                            original_value = EXCLUDED.original_value
                        """,
                        (
                            food_id,
                            nutrient_id,
                            normalize_number(original_value),
                            None if original_value is None else str(original_value),
                            "CIQUAL",
                            SOURCE_URL,
                            header,
                            dataset_version_id,
                        ),
                    )

                imported += 1

            cur.execute(
                "INSERT INTO import_logs (dataset_version_id, status, message) VALUES (%s, %s, %s)",
                (dataset_version_id, "success", f"Imported {imported} CIQUAL foods"),
            )

        conn.commit()
        print(json.dumps({"dataset_version_id": dataset_version_id, "imported_foods": imported}, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="Import CIQUAL XLSX into PostgreSQL.")
    parser.add_argument("xlsx", help="Path to official CIQUAL XLSX file")
    parser.add_argument("--version", default="2025", help="Dataset version label")
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL"), help="PostgreSQL connection URL")
    args = parser.parse_args()

    if not args.database_url:
        raise RuntimeError("DATABASE_URL is required")

    import_workbook(Path(args.xlsx), version=args.version, database_url=args.database_url)


if __name__ == "__main__":
    main()
